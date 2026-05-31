"""
shared/excel.py — Valorius
Convierte CSV de scraper a Excel formateado listo para revisión.
Uso: from shared.excel import csv_to_excel
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colores por estado ───────────────────────────────────────────────────────
FILLS = {
    "NUEVO":              PatternFill("solid", fgColor="C6EFCE"),  # verde
    "REVISAR":            PatternFill("solid", fgColor="FFEB9C"),  # amarillo
    "PROBABLE_DUPLICADO": PatternFill("solid", fgColor="FFCC99"),  # naranja
    "DUPLICADO_URL":      PatternFill("solid", fgColor="EFEFEF"),  # gris claro
}

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # azul oscuro Valorius
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Calibri", size=10)
ALERT_FONT    = Font(name="Calibri", size=10, color="C00000")  # rojo para alertas
THIN_BORDER   = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

# Columnas a mostrar y sus anchos (nombre_csv: ancho)
COLUMN_WIDTHS = {
    "estado_revision":        14,
    "alertas":                22,
    "fuente":                  9,
    "url":                    45,
    "titulo":                 35,
    "precio_usd":             12,
    "area_construccion":      14,
    "area_terreno":           12,
    "precio_m2":              10,
    "habitaciones":           11,
    "banos":                   8,
    "estacionamientos":       13,
    "tipo_inmueble":          12,
    "ubicacion_raw":          28,
    "zona_canonica_sugerida": 24,
    "zona_match_confianza":   14,
    "calidad_dato":           11,
    "proyecto_texto":         28,
    "fecha_scraping":         13,
    "match_motivo":           35,
    "zone_id_sugerido":       36,
}

# Columnas numéricas para formato
NUMERIC_COLS = {"precio_usd", "area_construccion", "area_terreno", "precio_m2",
                "habitaciones", "banos", "estacionamientos"}


def csv_to_excel(csv_path: Path | str, excel_path: Path | str = None) -> Path:
    csv_path   = Path(csv_path)
    excel_path = Path(excel_path) if excel_path else csv_path.with_suffix(".xlsx")

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader  = csv.DictReader(f)
        rows    = list(reader)
        headers = reader.fieldnames or []

    # Solo mostrar columnas definidas, en ese orden
    visible_cols = [c for c in COLUMN_WIDTHS if c in headers]
    # Agregar el resto al final por si hay columnas nuevas
    extra_cols   = [c for c in headers if c not in COLUMN_WIDTHS]
    all_cols     = visible_cols + extra_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Scraper"

    # ─── Encabezados ──────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell              = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").upper())
        cell.fill         = HEADER_FILL
        cell.font         = HEADER_FONT
        cell.alignment    = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ancho             = COLUMN_WIDTHS.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"

    # ─── Filas ────────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        estado = row.get("estado_revision", "")
        fill   = FILLS.get(estado)

        for col_idx, col_name in enumerate(all_cols, start=1):
            raw_val = row.get(col_name, "")

            # Convertir numéricos
            if col_name in NUMERIC_COLS and raw_val:
                try:
                    value = float(raw_val)
                    if col_name in {"habitaciones", "estacionamientos"}:
                        value = int(value)
                except ValueError:
                    value = raw_val
            else:
                value = raw_val

            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = ALERT_FONT if col_name == "alertas" and raw_val else DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if fill:
                cell.fill = fill

            # Formato numérico
            if col_name == "precio_usd" and isinstance(value, float):
                cell.number_format = '"$"#,##0'
            elif col_name in {"area_construccion", "area_terreno"} and isinstance(value, float):
                cell.number_format = '#,##0.0 "m²"'
            elif col_name == "precio_m2" and isinstance(value, float):
                cell.number_format = '"$"#,##0'

        ws.row_dimensions[row_idx].height = 16

    wb.save(excel_path)
    return excel_path
